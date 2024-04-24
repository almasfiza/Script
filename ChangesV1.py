import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from openpyxl.styles import PatternFill


import pandas as pd
import os

# Mapping for renaming extracted files
file_name_mapping_old = {
    "Bayer's Lake": "Bayers",
    "Cobequid": "Cobequid",
    "Dartmouth": "Dartmouth",
    "Halifax Infirmary": "Halifax",
    "Victoria General": "Victoria"
}
file_name_mapping_new = {
    "Bayer'S Lake Community Outpatient Center": "Bayers",
    "Cobequid Community Health Centre": "Cobequid",
    "Dartmouth General Hospital": "DG",
    "Halifax Infirmary Hospital": "HI",
    "Victoria General": "VG"
}

# Function to load data from Excel file
def load_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

# Function to compare changes between old and new data
def find_changes(old_data, new_data):
    old_numbers = {row[0] for row in old_data}  # Assuming "Number" is the first column
    new_numbers = {row[0] for row in new_data}  # Assuming "Number" is the first column

    # if the record is not present in new numbers, then add to removed 
    # otherwise add it to new record list
    removed_records = [row for row in old_data if row[0] not in new_numbers]
    new_records = [row for row in new_data if row[0] not in old_numbers]

    return new_records, removed_records

# Function to highlight changes and save them to a new Excel file
def highlight_changes(new_records, removed_records, old_data, new_data, output_path):
    wb = openpyxl.Workbook()

    # Deleting the default sheet and adding two new worksheets of New and Old records.
    del wb["Sheet"]
    new_ws = wb.create_sheet(title="Added")
    removed_ws = wb.create_sheet(title="Removed")

    # Filter new records based on existence in new data and absence in old data
    filtered_new_records = [record for record in new_records if record in new_data and record not in old_data]

    # Filter removed records based on existence in old data and absence in new data
    filtered_removed_records = [record for record in removed_records if record in old_data and record not in new_data]

    # Write new records
    for record in filtered_new_records:
        new_ws.append(record)
   
    # Write removed records
    for record in filtered_removed_records:
        removed_ws.append(record)

    # Save the workbook
    wb.save(output_path)

# Function to extract old data and automatically load it into the GUI
def extract_old_data_and_load_gui(app):
    input_folder_path = filedialog.askdirectory()  # Select a folder containing Excel files
    if input_folder_path:
        try:
            # Determine the directory of the script
            script_dir = os.path.dirname(__file__)

            # Create the output directory for extracted old data
            extracted_files_folder = os.path.join(script_dir, "EXTRACTED OLD DATA")
            os.makedirs(extracted_files_folder, exist_ok=True)

            # Iterate over each file in the selected folder
            for filename in os.listdir(input_folder_path):
                if filename.endswith(".xlsx"):
                    input_file_path = os.path.join(input_folder_path, filename)
                    
                    # Extract facility name from the filename
                    for facility_name, new_name in file_name_mapping_old.items():
                        if facility_name.lower() in filename.lower():
                            extracted_file_path = os.path.join(extracted_files_folder, f"extracted_{new_name}.xlsx")

                            # Extract old data from the current file
                            save_worksheet_as_excel(input_file_path, "All Facility Data", extracted_file_path)
                            break  # Stop searching once a match is found

        except Exception as e:
            messagebox.showerror("Error", str(e))

# Function to extract new data and display extracted files for selection

def filter_and_store_facilities(input_file, output_folder):
    # Read the Excel file
    df = pd.read_excel(input_file)

    # Filter data for Centennial, Dickson, Veterans', and Victoria Building
    victoria_data = df[df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterens' Memorial Building", 'Victoria Building'])]

    # Create the output folder
    os.makedirs(output_folder, exist_ok=True)

    # Create an Excel writer object for Victoria Building
    output_file_victoria = os.path.join(output_folder, "Victoria General.xlsx")
    with pd.ExcelWriter(output_file_victoria, engine='openpyxl') as writer_victoria:
        # Store data of Centennial, Dickson, Veterans', and Victoria Building in the same worksheet
        victoria_data.to_excel(writer_victoria, sheet_name='Victoria Building', index=False)

    # Store data for other facilities in separate Excel files
    other_facilities_data = df[~df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterens' Memorial Building", 'Victoria Building'])]
    for facility_name, data in other_facilities_data.groupby('Facility Name'):
        output_file = os.path.join(output_folder, f"{facility_name}.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, index=False)


# Function to extract a specific worksheet from an Excel file
def save_worksheet_as_excel(input_file, sheet_name, output_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Select the desired worksheet
    ws = wb[sheet_name]
    
    # Create a new workbook and copy the contents of the selected worksheet
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Copy cell values and formatting
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)
    
    # Save the new workbook
    new_wb.save(output_file)


def extract_new_data_and_load_gui(app):
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if input_file_path:
        try:
            # Determine the directory of the script
            script_dir = os.path.dirname(__file__)

            # Create the output directory for new extracted data
            extracted_files_folder = os.path.join(script_dir, "EXTRACTED NEW DATA")
            os.makedirs(extracted_files_folder, exist_ok=True)
            
            # Extract new data and store in the output directory
            filter_and_store_facilities(input_file_path, extracted_files_folder)
            
            # Display extracted files for selection
            selected_file = select_file_from_folder(extracted_files_folder)
            if selected_file:
                app.new_file_path = selected_file

                # Get the base filename without extension
                base_filename = os.path.splitext(os.path.basename(selected_file))[0]

                # Iterate over the mapping to find the corresponding display name
                for facility_name, display_name in file_name_mapping_new.items():
                    if display_name in base_filename:
                        app.new_file_label.config(text=f"{facility_name}.xlsx")
                        break  # Stop after finding the matching display name
        
        except Exception as e:
            messagebox.showerror("Error", str(e))


# Function to filter and store facilities into separate Excel files
def filter_and_store_facilities(input_file, output_folder):
    # Read the Excel file
    df = pd.read_excel(input_file)

    # Filter data for Centennial, Dickson, Veterans', and Victoria Building
    victoria_data = df[df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterans' Memorial Building", 'Victoria Building', 'Bethune Building', 'Abbie J. Lane'])]

    # Filter data for other facilities
    other_facilities_data = df[~df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterans' Memorial Building", 'Victoria Building', 'Bethune Building', 'Abbie J. Lane'])]

    # Create the output folder
    os.makedirs(output_folder, exist_ok=True)

    # Create an Excel writer object for Victoria Building
    output_file_victoria = os.path.join(output_folder, "Victoria General.xlsx")
    with pd.ExcelWriter(output_file_victoria, engine='openpyxl') as writer_victoria:
        # Store data of Centennial, Dickson, Veterans', and Victoria Building in the same worksheet
        victoria_data.to_excel(writer_victoria, index=False)

    # Store data for other facilities in separate Excel files
    for facility_name, data in other_facilities_data.groupby('Facility Name'):
        output_file = os.path.join(output_folder, f"{facility_name}.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, index=False)



# Function to display extracted files for selection
def select_file_from_folder(folder_path):
    files = os.listdir(folder_path)
    selected_file = filedialog.askopenfilename(initialdir=folder_path, title="Select File", filetypes=[("Excel files", "*.xlsx")])
    return selected_file

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Comparator")
        self.geometry("400x300")
        
        # Initialize file paths
        self.old_file_path = ""
        self.new_file_path = ""
        
        # Create labels
        self.label1 = tk.Label(self, text="Select Old File:")
        self.label1.grid(row=0, column=0, padx=10, pady=5)
        
        self.label2 = tk.Label(self, text="Select New File:")
        self.label2.grid(row=1, column=0, padx=10, pady=5)
        
        # Create labels to display selected file names
        self.old_file_label = tk.Label(self, text="")
        self.old_file_label.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        
        self.new_file_label = tk.Label(self, text="")
        self.new_file_label.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        
        # Create buttons
        self.old_file_button = tk.Button(self, text="Browse", command=self.select_old_file)
        self.old_file_button.grid(row=0, column=2, padx=10, pady=5)

        self.old_file_button.configure(bg="green")  # Set initial color to green
        
        self.new_file_button = tk.Button(self, text="Browse", command=self.select_new_file)
        self.new_file_button.grid(row=1, column=2, padx=10, pady=5)

        self.new_file_button.configure(bg="green")  # Set initial color to green

        self.extract_old_data_button = tk.Button(self, text="Extract Old Data", command=self.extract_old_data)
        self.extract_old_data_button.grid(row=2, column=0, columnspan=3, padx=10, pady=5)

        self.extract_old_data_button.configure(bg="green")  # Set initial color to green
        
        self.extract_new_data_button = tk.Button(self, text="Extract New Data", command=self.extract_new_data)
        self.extract_new_data_button.grid(row=3, column=0, columnspan=3, padx=10, pady=5)

        self.extract_new_data_button.configure(bg="green")  # Set initial color to green

        self.compare_button = tk.Button(self, text="Compare", command=self.compare_files)
        self.compare_button.grid(row=4, column=0, columnspan=3, padx=10, pady=5)
        
        self.compare_button.configure(bg="green")  # Set initial color to green
        
         # Create labels and buttons for selecting folders
        self.label3 = tk.Label(self, text="Select Old Folder:")
        self.label3.grid(row=5, column=0, padx=10, pady=5)
        
        self.old_folder_label = tk.Label(self, text="")
        self.old_folder_label.grid(row=5, column=1, padx=10, pady=5, sticky="w")
        
        self.old_folder_button = tk.Button(self, text="Browse", command=self.select_old_folder)
        self.old_folder_button.grid(row=5, column=2, padx=10, pady=5)

        self.label4 = tk.Label(self, text="Select New Folder:")
        self.label4.grid(row=6, column=0, padx=10, pady=5)
        
        self.new_folder_label = tk.Label(self, text="")
        self.new_folder_label.grid(row=6, column=1, padx=10, pady=5, sticky="w")
        
        self.new_folder_button = tk.Button(self, text="Browse", command=self.select_new_folder)
        self.new_folder_button.grid(row=6, column=2, padx=10, pady=5)

        # Create button for comparing all files
        self.compare_all_button = tk.Button(self, text="Compare All", command=self.compare_all_files)
        self.compare_all_button.grid(row=7, column=0, columnspan=3, padx=10, pady=5)


        
    def select_old_folder(self):
        self.old_folder_path = filedialog.askdirectory()
        if self.old_folder_path:
            self.old_folder_label.config(text=self.old_folder_path)
        
    def select_new_folder(self):
        self.new_folder_path = filedialog.askdirectory()
        if self.new_folder_path:
            self.new_folder_label.config(text=self.new_folder_path)

    def select_old_file(self):
        self.old_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.old_file_label.config(text=os.path.basename(self.old_file_path))
        
    def select_new_file(self):
        self.new_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.new_file_label.config(text=os.path.basename(self.new_file_path))
        
    def extract_old_data(self):
        self.extract_old_data_button.config(text="Extracting...", bg="red")
        self.update_idletasks()  # Update the display
        extract_old_data_and_load_gui(self)
        self.extract_old_data_button.config(text="Extract Old Data", bg="green")

    def extract_new_data(self):
        self.extract_new_data_button.config(text="Extracting...", bg="red")
        self.update_idletasks()  # Update the display
        extract_new_data_and_load_gui(self)
        self.extract_new_data_button.config(text="Extract New Data", bg="green")
        
    def compare_files(self):
        self.compare_button.config(text="Comparing...", bg="red")
        self.update_idletasks()  # Update the display
        if not self.old_file_path or not self.new_file_path:
            messagebox.showerror("Error", "Please select both old and new files.")
            self.compare_button.config(text="Compare", bg="green")
            return
        
        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if output_file_path:
            try:
                old_data = load_data(self.old_file_path)
                new_data = load_data(self.new_file_path)

                new_records, removed_records = find_changes(old_data, new_data)

                highlight_changes(new_records, removed_records, old_data, new_data, output_file_path)

                messagebox.showinfo("Success", "Comparison completed! Output file saved.")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        self.compare_button.config(text="Compare", bg="green")


    def compare_all_files(self):
        if not self.old_folder_path or not self.new_folder_path:
            messagebox.showerror("Error", "Please select both old and new folders.")
            return
        
        # Create the 'RESULT ALL' folder if it does not exist in the script's root directory
        result_folder_path = os.path.join(os.path.dirname(__file__), "RESULT ALL")
        os.makedirs(result_folder_path, exist_ok=True)
        
        # Mapping of extracted names to full names
        file_mapping = {
            "Bayers": "Bayer'S Lake Community Outpatient Center",
            "Cobequid": "Cobequid Community Health Centre",
            "Dartmouth": "Dartmouth General Hospital",
            "Halifax": "Halifax Infirmary Hospital",
            "Victoria": "Victoria General"
        }
        
        # Iterate over the file mapping and compare files
        for extracted_name, full_name in file_mapping.items():
            old_file_path = os.path.join(self.old_folder_path, f"extracted_{extracted_name}.xlsx")
            new_file_path = os.path.join(self.new_folder_path, f"{full_name}.xlsx")
            result_file_name = f"{extracted_name}_Res.xlsx"
            result_file_path = os.path.join(result_folder_path, result_file_name)
            
            try:
                old_data = load_data(old_file_path)
                new_data = load_data(new_file_path)
                
                old_base_name = f"extracted_{extracted_name}"
                
                # Perform comparison using the same logic as individual file comparison
                new_records, removed_records = find_changes(old_data, new_data)
                highlight_changes(new_records, removed_records, old_data, new_data, result_file_path)
                
                messagebox.showinfo("Success", f"Comparison completed for {old_base_name} and {full_name}.")
            except Exception as e:
                messagebox.showerror("Error", str(e))


            
if __name__ == "__main__":
    app = Application()
    app.mainloop()
