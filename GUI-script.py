import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from openpyxl.styles import PatternFill


import pandas as pd
import os

# Function to load data from Excel file
def load_data(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

# Function to compare changes between old and new data
def find_changes(old_data, new_data):
    old_numbers = {row[0] for row in old_data}  # Assuming "Number" is the first column
    new_numbers = {row[0] for row in new_data}  # Assuming "Number" is the first column

    # if the record is not present in new numbers, then add to removed 
    # otherwise add it to new record list
    removed_records = [row for row in old_data if row[0] not in new_numbers]
    new_records = [row for row in new_data if row[0] not in old_numbers]

    return new_records, removed_records

# Function to highlight changes and save them to a new Excel file
def highlight_changes(new_records, removed_records, old_data, new_data, output_path):
    wb = openpyxl.Workbook()

    # Deleting the default sheet and adding two new worksheets of New and Old records.
    del wb["Sheet"]
    new_ws = wb.create_sheet(title="Added")
    removed_ws = wb.create_sheet(title="Removed")

    # Filter new records based on existence in new data and absence in old data
    filtered_new_records = [record for record in new_records if record in new_data and record not in old_data]

    # Filter removed records based on existence in old data and absence in new data
    filtered_removed_records = [record for record in removed_records if record in old_data and record not in new_data]

    # Write new records
    for record in filtered_new_records:
        new_ws.append(record)
   
    # Write removed records
    for record in filtered_removed_records:
        removed_ws.append(record)

    # Save the workbook
    wb.save(output_path)

# Function to extract old data and automatically load it into the GUI
def extract_old_data_and_load_gui(app):
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if input_file_path:
        try:
            # Extract old data
            extracted_file_path = "temp_extracted_old_data.xlsx"  # Temporary file path
            save_worksheet_as_excel(input_file_path, "All Facility Data", extracted_file_path)
            
            # Load extracted data into the GUI
            app.old_file_path = extracted_file_path
            app.old_file_label.config(text=os.path.basename(extracted_file_path))
        except Exception as e:
            messagebox.showerror("Error", str(e))

# Function to extract new data and display extracted files for selection

def filter_and_store_facilities(input_file, output_folder):
    # Read the Excel file
    df = pd.read_excel(input_file)

    # Filter data for Centennial, Dickson, Veterans', and Victoria Building
    victoria_data = df[df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterens' Memorial Building", 'Victoria Building'])]

    # Create the output folder
    os.makedirs(output_folder, exist_ok=True)

    # Create an Excel writer object for Victoria Building
    output_file_victoria = os.path.join(output_folder, "Victoria_Building.xlsx")
    with pd.ExcelWriter(output_file_victoria, engine='openpyxl') as writer_victoria:
        # Store data of Centennial, Dickson, Veterans', and Victoria Building in the same worksheet
        victoria_data.to_excel(writer_victoria, sheet_name='Victoria Building', index=False)

    # Store data for other facilities in separate Excel files
    other_facilities_data = df[~df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterens' Memorial Building", 'Victoria Building'])]
    for facility_name, data in other_facilities_data.groupby('Facility Name'):
        output_file = os.path.join(output_folder, f"{facility_name}.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, index=False)


# Function to extract a specific worksheet from an Excel file
def save_worksheet_as_excel(input_file, sheet_name, output_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Select the desired worksheet
    ws = wb[sheet_name]
    
    # Create a new workbook and copy the contents of the selected worksheet
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    
    # Copy cell values and formatting
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)
    
    # Save the new workbook
    new_wb.save(output_file)


# Function to extract new data and display extracted files for selection
def extract_new_data_and_load_gui(app):
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if input_file_path:
        try:
            # Extract new data
            extracted_files_folder = "temp_extracted_new_data"  # Temporary folder path
            
            # Create the output directory if it doesn't exist
            os.makedirs(extracted_files_folder, exist_ok=True)
            
            filter_and_store_facilities(input_file_path, extracted_files_folder)
            
            # Display extracted files for selection
            selected_file = select_file_from_folder(extracted_files_folder)
            if selected_file:
                app.new_file_path = selected_file
                app.new_file_label.config(text=os.path.basename(selected_file))
        except Exception as e:
            messagebox.showerror("Error", str(e))
# Function to filter and store facilities into separate Excel files
def filter_and_store_facilities(input_file, output_folder):
    # Read the Excel file
    df = pd.read_excel(input_file)

    # Filter data for Centennial, Dickson, Veterans', and Victoria Building
    victoria_data = df[df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterans' Memorial Building", 'Victoria Building'])]

    # Filter data for other facilities
    other_facilities_data = df[~df['Facility Name'].isin(['Centennial Building', 'Dickson Building', "Veterans' Memorial Building", 'Victoria Building'])]

    # Create the output folder
    os.makedirs(output_folder, exist_ok=True)

    # Create an Excel writer object for Victoria Building
    output_file_victoria = os.path.join(output_folder, "Victoria_Building.xlsx")
    with pd.ExcelWriter(output_file_victoria, engine='openpyxl') as writer_victoria:
        # Store data of Centennial, Dickson, Veterans', and Victoria Building in the same worksheet
        victoria_data.to_excel(writer_victoria, index=False)

    # Store data for other facilities in separate Excel files
    for facility_name, data in other_facilities_data.groupby('Facility Name'):
        output_file = os.path.join(output_folder, f"{facility_name}.xlsx")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            data.to_excel(writer, index=False)



# Function to display extracted files for selection
def select_file_from_folder(folder_path):
    files = os.listdir(folder_path)
    selected_file = filedialog.askopenfilename(initialdir=folder_path, title="Select File", filetypes=[("Excel files", "*.xlsx")])
    return selected_file

class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Comparator")
        self.geometry("400x300")
        
        # Initialize file paths
        self.old_file_path = ""
        self.new_file_path = ""
        
        # Create labels
        self.label1 = tk.Label(self, text="Select Old File:")
        self.label1.grid(row=0, column=0, padx=10, pady=5)
        
        self.label2 = tk.Label(self, text="Select New File:")
        self.label2.grid(row=1, column=0, padx=10, pady=5)
        
        # Create labels to display selected file names
        self.old_file_label = tk.Label(self, text="")
        self.old_file_label.grid(row=0, column=1, padx=10, pady=5, sticky="w")
        
        self.new_file_label = tk.Label(self, text="")
        self.new_file_label.grid(row=1, column=1, padx=10, pady=5, sticky="w")
        
        # Create buttons
        self.old_file_button = tk.Button(self, text="Browse", command=self.select_old_file)
        self.old_file_button.grid(row=0, column=2, padx=10, pady=5)
        
        self.new_file_button = tk.Button(self, text="Browse", command=self.select_new_file)
        self.new_file_button.grid(row=1, column=2, padx=10, pady=5)

        self.extract_old_data_button = tk.Button(self, text="Extract Old Data", command=lambda: extract_old_data_and_load_gui(self))
        self.extract_old_data_button.grid(row=2, column=0, columnspan=3, padx=10, pady=5)
        
        self.extract_new_data_button = tk.Button(self, text="Extract New Data", command=lambda: extract_new_data_and_load_gui(self))
        self.extract_new_data_button.grid(row=3, column=0, columnspan=3, padx=10, pady=5)

        self.compare_button = tk.Button(self, text="Compare", command=self.compare_files)
        self.compare_button.grid(row=4, column=0, columnspan=3, padx=10, pady=5)
        
    def select_old_file(self):
        self.old_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.old_file_label.config(text=os.path.basename(self.old_file_path))
        
    def select_new_file(self):
        self.new_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.new_file_label.config(text=os.path.basename(self.new_file_path))
        
    def compare_files(self):
        if not self.old_file_path or not self.new_file_path:
            messagebox.showerror("Error", "Please select both old and new files.")
            return
        
        output_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if output_file_path:
            try:
                old_data = load_data(self.old_file_path)
                new_data = load_data(self.new_file_path)

                new_records, removed_records = find_changes(old_data, new_data)

                highlight_changes(new_records, removed_records, old_data, new_data, output_file_path)

                messagebox.showinfo("Success", "Comparison completed! Output file saved.")
            except Exception as e:
                messagebox.showerror("Error", str(e))
        
if __name__ == "__main__":
    app = Application()
    app.mainloop()
